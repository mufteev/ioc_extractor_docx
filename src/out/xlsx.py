
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ioc.normalizer import IoCType
from ioc.extractor import ExtractionResult

# ============================================================================
# Экспорт результатов в Excel
# ============================================================================



def export_to_excel(
    results: dict[str, ExtractionResult],
    output_path: str | Path,
    include_context: bool = True,
    sheet_name_max_length: int = 31  # Ограничение Excel на длину имени листа
) -> Path:
    """
    Экспортирует результаты извлечения IoC в Excel-файл.
    
    Каждый исходный файл сохраняется на отдельном листе с колонками:
    - value: нормализованное значение IoC
    - type: тип индикатора (HASH_SHA256, URL, IP_ADDRESS и т.д.)
    - original: исходное значение до нормализации
    - defanged: был ли индикатор "обезврежен" (True/False)
    - context: контекст, в котором найден индикатор
    
    Args:
        results: Словарь {путь_к_файлу: ExtractionResult} из extract_from_files()
        output_path: Путь для сохранения Excel-файла
        include_context: Включать ли колонку context (может быть длинной)
        sheet_name_max_length: Максимальная длина имени листа (по умолчанию 31 — лимит Excel)
        
    Returns:
        Path к созданному файлу
        
    Raises:
        ImportError: Если openpyxl не установлен
        
    Пример:
        results = extract_iocs_from_files(["report1.docx", "report2.docx"])
        export_to_excel(results, "iocs_export.xlsx")
    """
    output_path = Path(output_path)
    
    # Создаём новую книгу Excel
    wb = Workbook()
    
    # Удаляем дефолтный лист, если будем создавать свои
    if isinstance(wb.active, Worksheet):
      default_sheet: Worksheet = wb.active
      # Теперь можно безопасно работать с default_sheet
      default_sheet['A1'] = 'Данные'
    else:
        # Обрабатываем неожиданную ситуацию
        raise ValueError("Активный лист не является обычным рабочим листом")
    
    # Стили для заголовков
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Стили для данных
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(vertical='top', wrap_text=False)
    data_alignment_wrap = Alignment(vertical='top', wrap_text=True)
    
    # Стиль для дефангированных значений (подсветка)
    defanged_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    # Границы
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Определяем колонки
    if include_context:
        headers = ['value', 'type', 'original', 'defanged', 'rule_extracted', 'context']
        column_widths = [60, 15, 60, 10, 20, 50]  # Ширина колонок в символах
    else:
        headers = ['value', 'type', 'original', 'defanged', 'rule_extracted']
        column_widths = [60, 15, 60, 10, 20]
    # Отслеживаем использованные имена листов для уникальности
    used_sheet_names: set[str] = set()
    
    def make_unique_sheet_name(filepath: str) -> str:
        """
        Создаёт уникальное и валидное имя листа из пути к файлу.
        
        Excel имеет ограничения:
        - Максимум 31 символ
        - Запрещены символы: \\ / * ? : [ ]
        """
        # Извлекаем имя файла без расширения
        name = Path(filepath).stem
        
        # Удаляем запрещённые символы
        forbidden_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in forbidden_chars:
            name = name.replace(char, '_')
        
        # Обрезаем до максимальной длины (оставляем место для суффикса)
        base_name = name[:sheet_name_max_length - 4]  # -4 для возможного суффикса "_99"
        
        # Делаем имя уникальным
        final_name = base_name
        counter = 1
        while final_name.lower() in used_sheet_names:
            final_name = f"{base_name}_{counter}"
            counter += 1
        
        used_sheet_names.add(final_name.lower())
        return final_name
    
    # Флаг: был ли использован дефолтный лист
    first_sheet_used = False
    
    # Создаём лист для каждого файла
    for filepath, result in results.items():
        sheet_name = make_unique_sheet_name(filepath)

        
        # Используем дефолтный лист для первого файла, потом создаём новые
        if not first_sheet_used:
            ws = default_sheet
            ws.title = sheet_name
            first_sheet_used = True
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Устанавливаем ширину колонок
        for col_idx, width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
        
        # Записываем данные
        for row_idx, ioc in enumerate(result.iocs, start=2):
            # value
            cell_value = ws.cell(row=row_idx, column=1, value=ioc.value)
            cell_value.font = data_font
            cell_value.alignment = data_alignment
            cell_value.border = thin_border
            
            # type
            cell_type = ws.cell(row=row_idx, column=2, value=ioc.ioc_type.name)
            cell_type.font = data_font
            cell_type.alignment = Alignment(horizontal='center', vertical='top')
            cell_type.border = thin_border
            
            # original
            cell_original = ws.cell(row=row_idx, column=3, value=ioc.original_value)
            cell_original.font = data_font
            cell_original.alignment = data_alignment
            cell_original.border = thin_border
            
            # defanged
            cell_defanged = ws.cell(row=row_idx, column=4, value=ioc.defanged)
            cell_defanged.font = data_font
            cell_defanged.alignment = Alignment(horizontal='center', vertical='top')
            cell_defanged.border = thin_border
            
            # Подсвечиваем строку, если IoC был дефангирован
            if ioc.defanged:
                cell_value.fill = defanged_fill
                cell_type.fill = defanged_fill
                cell_original.fill = defanged_fill
                cell_defanged.fill = defanged_fill
             
            # rule_extracted
            cell_rule = ws.cell(row=row_idx, column=5, value=ioc.rule_extracted)
            cell_rule.font = data_font
            cell_rule.alignment = Alignment(horizontal='center', vertical='top')
            cell_rule.border = thin_border

            # context (опционально)
            if include_context:
                # Обрезаем слишком длинный контекст
                context_text = ioc.source_context
                if len(context_text) > 500:
                    context_text = context_text[:497] + "..."
                
                cell_context = ws.cell(row=row_idx, column=6, value=context_text)
                cell_context.font = data_font
                cell_context.alignment = data_alignment_wrap
                cell_context.border = thin_border
                
                if ioc.defanged:
                    cell_context.fill = defanged_fill
        
        # Закрепляем заголовок (freeze panes)
        ws.freeze_panes = 'A2'
        
        # Добавляем автофильтр
        if result.iocs:
            last_col = get_column_letter(len(headers))
            last_row = len(result.iocs) + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    
    # Если не было файлов с результатами, создаём пустой лист с информацией
    if not first_sheet_used:
        ws = default_sheet
        ws.title = "No Results"
        ws['A1'] = "Индикаторы компрометации не найдены"
        ws['A1'].font = Font(name='Arial', size=12, italic=True)
    
    # Добавляем сводный лист в начало
    summary_sheet = wb.create_sheet(title="Summary", index=0)
    _create_summary_sheet(summary_sheet, results, headers, header_font, header_fill, 
                          header_alignment, data_font, thin_border)
    
    # Сохраняем файл
    wb.save(output_path)
    
    return output_path



def _create_summary_sheet(
    ws,
    results: dict[str, ExtractionResult],
    headers: list[str],
    header_font,
    header_fill,
    header_alignment,
    data_font,
    thin_border
) -> None:
    """
    Создаёт сводный лист со статистикой по всем файлам.
    
    Это вспомогательная функция для export_to_excel().
    """
    
    # Локальные стили для сводки
    bold_font = Font(name='Arial', bold=True, size=11)
    title_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
    
    # Заголовок сводки
    ws['A1'] = "IoC Extraction Summary"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Заголовки таблицы статистики
    stat_headers = ['File', 'Total IoCs', 'Hashes', 'URLs', 'IPs', 'Other', 'Errors']
    for col_idx, header in enumerate(stat_headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Данные по каждому файлу
    row_idx = 4
    total_iocs = 0
    total_hashes = 0
    total_urls = 0
    total_ips = 0
    total_other = 0
    total_errors = 0
    
    for filepath, result in results.items():
        # Подсчитываем статистику
        hashes = len([i for i in result.iocs if i.ioc_type.name.startswith('HASH_')])
        urls = len(result.by_type(IoCType.URL))
        ips = len(result.by_type(IoCType.IP_ADDRESS))
        other = len(result) - hashes - urls - ips
        errors = len(result.errors)
        
        # Накапливаем итоги
        total_iocs += len(result)
        total_hashes += hashes
        total_urls += urls
        total_ips += ips
        total_other += other
        total_errors += errors
        
        # Записываем строку
        filename = Path(filepath).name
        row_data = [filename, len(result), hashes, urls, ips, other, errors]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal='center')
        
        row_idx += 1
    
    # Итоговая строка
    total_cell = ws.cell(row=row_idx, column=1, value="TOTAL")
    total_cell.font = bold_font
    total_cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    totals = [total_iocs, total_hashes, total_urls, total_ips, total_other, total_errors]
    for col_idx, value in enumerate(totals, start=2):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        cell.border = thin_border
    
    # Ширина колонок
    column_widths = [40, 12, 12, 12, 12, 12, 12]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Закрепляем заголовки
    ws.freeze_panes = 'A4'
